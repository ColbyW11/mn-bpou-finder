"""
Generate BPOUMap.geojson from voting precinct shapefile + precinct allocation Excel.

Usage: python3 generate_bpou_geojson.py
"""

import json
import warnings
import geopandas as gpd
import openpyxl

SHAPEFILE = "shp_bdry_votingdistricts/bdry_votingdistricts.shp"
ALLOCATION_XLSX = "25_26_CD_Bpou_PctAlloc_Prot_260113b(5).xlsx"
OUTPUT = "BPOUMap.geojson"

# Name transformations to match existing BPOU_NAME convention
NAME_OVERRIDES = {
    "Minneapolis BPOU": "Minneapolis Republicans",
    "St Louis County": "Saint Louis County Republicans",
    "Lac Qui Parle County": "Lac qui Parle County Republicans",
    "HD 03B": "HD3B Republicans",
}


def transform_bpou_name(excel_name):
    """Convert Excel BPOU name to the BPOU_NAME format used in GeoJSON/contacts."""
    if excel_name in NAME_OVERRIDES:
        return NAME_OVERRIDES[excel_name]
    # HD/SD: remove space between prefix and number (e.g. "HD 27A" -> "HD27A")
    if excel_name.startswith("HD ") or excel_name.startswith("SD "):
        return excel_name.replace(" ", "") + " Republicans"
    # County BPOUs and everything else
    return excel_name + " Republicans"


def read_allocation(path):
    """Read Excel allocation file, return dict of CoPct -> BPOU_NAME."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    mapping = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # skip header
            continue
        copct = str(row[2]).zfill(6) if row[2] else None
        bpou_name = row[1]
        if copct and bpou_name:
            mapping[copct] = transform_bpou_name(bpou_name)
    wb.close()
    return mapping


def main():
    # Read shapefile
    print(f"Reading shapefile: {SHAPEFILE}")
    gdf = gpd.read_file(SHAPEFILE)
    print(f"  {len(gdf)} precincts loaded (CRS: {gdf.crs})")

    # Build CoPct key from shapefile
    gdf["CoPct"] = (
        gdf["COUNTYCODE"].astype(str).str.zfill(2)
        + gdf["PCTCODE"].astype(str).str.zfill(4)
    )

    # Read allocation
    print(f"Reading allocation: {ALLOCATION_XLSX}")
    alloc = read_allocation(ALLOCATION_XLSX)
    print(f"  {len(alloc)} precinct allocations loaded")

    # Join: add BPOU_NAME to shapefile based on CoPct
    gdf["BPOU_NAME"] = gdf["CoPct"].map(alloc)

    # For unmatched precincts, infer BPOU from house district using the most common
    # BPOU assignment for that HD within the same county
    unmatched = gdf[gdf["BPOU_NAME"].isna()]
    if len(unmatched) > 0:
        # Build HD -> BPOU lookup from matched precincts in the same county
        matched_so_far = gdf[gdf["BPOU_NAME"].notna()]
        hd_county_bpou = (
            matched_so_far.groupby(["COUNTYCODE", "MNLEGDIST"])["BPOU_NAME"]
            .agg(lambda x: x.mode().iloc[0])
            .to_dict()
        )
        inferred = 0
        still_unmatched = 0
        for idx, row in unmatched.iterrows():
            key = (row["COUNTYCODE"], row["MNLEGDIST"])
            if key in hd_county_bpou:
                gdf.at[idx, "BPOU_NAME"] = hd_county_bpou[key]
                inferred += 1
                print(f"  Inferred: CoPct={row['CoPct']} {row['PCTNAME']} -> {hd_county_bpou[key]}")
            else:
                still_unmatched += 1
                print(f"  WARNING: No match for CoPct={row['CoPct']} {row['PCTNAME']} ({row['COUNTYNAME']})")
        print(f"\n{inferred} precincts inferred from house district, {still_unmatched} still unmatched")

    # Also check for Excel entries with no shapefile match
    shp_copcts = set(gdf["CoPct"])
    excel_only = {k: v for k, v in alloc.items() if k not in shp_copcts}
    if excel_only:
        print(f"\nNOTE: {len(excel_only)} Excel precincts have no shapefile geometry (skipped):")
        for copct, bpou in sorted(excel_only.items()):
            print(f"  CoPct={copct} -> {bpou}")

    # Filter to matched precincts only
    matched = gdf[gdf["BPOU_NAME"].notna()].copy()
    print(f"\n{len(matched)} precincts matched, dissolving into BPOUs...")

    # Dissolve by BPOU_NAME (merge precinct geometries)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        dissolved = matched.dissolve(by="BPOU_NAME").reset_index()

    print(f"  {len(dissolved)} BPOUs created")

    # Keep only BPOU_NAME and geometry
    dissolved = dissolved[["BPOU_NAME", "geometry"]]

    # Reproject to WGS84 (EPSG:4326)
    dissolved = dissolved.to_crs(epsg=4326)

    # Export as GeoJSON
    print(f"Writing {OUTPUT}...")
    dissolved.to_file(OUTPUT, driver="GeoJSON")

    # Verify output
    with open(OUTPUT) as f:
        data = json.load(f)
    names = sorted([f["properties"]["BPOU_NAME"] for f in data["features"]])
    print(f"\nOutput: {len(names)} BPOU features")
    for name in names:
        print(f"  {name}")

    # Cross-check with contacts
    try:
        with open("bpouContacts.json") as f:
            contacts = json.load(f)
        geojson_names = set(names)
        contact_names = set(contacts.keys())
        missing_contacts = geojson_names - contact_names
        extra_contacts = contact_names - geojson_names
        if missing_contacts:
            print(f"\nWARNING: BPOUs in GeoJSON but NOT in contacts:")
            for n in sorted(missing_contacts):
                print(f"  {n}")
        if extra_contacts:
            print(f"\nNOTE: BPOUs in contacts but NOT in GeoJSON:")
            for n in sorted(extra_contacts):
                print(f"  {n}")
        if not missing_contacts and not extra_contacts:
            print("\nAll BPOU names match between GeoJSON and contacts.")
    except FileNotFoundError:
        pass

    print("\nDone!")


if __name__ == "__main__":
    main()
